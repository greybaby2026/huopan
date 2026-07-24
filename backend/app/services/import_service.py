"""Excel 导入服务: 解析 Excel 批量导入产品"""
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from app.core.config import settings


def generate_import_template() -> bytes:
    """生成产品导入模板 Excel, 返回 bytes"""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "产品导入"

    headers = [
        "款号", "名称", "品类", "颜色", "花型", "季节",
        "风格", "面料", "尺码", "成本价", "零售价", "库存", "状态", "备注"
    ]

    # 表头样式
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # 示例行
    example = ["LJ-SS001", "修身商务衬衫", "衬衫", "白色", "纯色", "春",
               "商务", "100%棉", "S,M,L,XL", 45, 129, 200, "active", "春季主推"]
    for col, v in enumerate(example, 1):
        ws.cell(row=2, column=col, value=v)

    # 列宽
    col_widths = [14, 18, 10, 8, 8, 8, 8, 14, 14, 10, 10, 8, 10, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def parse_import_excel(file_bytes: bytes) -> list[dict]:
    """
    解析产品导入 Excel, 返回产品字典列表.
    列顺序: 款号/名称/品类/颜色/花型/季节/风格/面料/尺码/成本价/零售价/库存/状态/备注
    """
    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=2, values_only=True))  # 跳过表头
    products = []

    for row in rows:
        if not row or not row[0]:
            continue  # 空行或无款号

        sku_code = str(row[0]).strip()
        if not sku_code:
            continue

        def safe_str(val, default=""):
            if val is None:
                return default
            return str(val).strip()

        def safe_float(val, default=0):
            try:
                return float(val) if val is not None else default
            except (ValueError, TypeError):
                return default

        def safe_int(val, default=0):
            try:
                return int(val) if val is not None else default
            except (ValueError, TypeError):
                return default

        product = {
            "sku_code": sku_code,
            "name": safe_str(row[1]) if len(row) > 1 else "",
            "category": safe_str(row[2]) or None if len(row) > 2 else None,
            "color": safe_str(row[3]) or None if len(row) > 3 else None,
            "pattern": safe_str(row[4]) or None if len(row) > 4 else None,
            "season": safe_str(row[5]) or None if len(row) > 5 else None,
            "style": safe_str(row[6]) or None if len(row) > 6 else None,
            "fabric": safe_str(row[7]) or None if len(row) > 7 else None,
            "size_range": safe_str(row[8]) or None if len(row) > 8 else None,
            "cost_price": safe_float(row[9]) if len(row) > 9 else 0,
            "retail_price": safe_float(row[10]) if len(row) > 10 else 0,
            "stock": safe_int(row[11]) if len(row) > 11 else 0,
            "status": safe_str(row[12], "draft") or "draft" if len(row) > 12 else "draft",
            "note": safe_str(row[13]) or None if len(row) > 13 else None,
        }

        # 名称必填
        if not product["name"]:
            product["name"] = sku_code

        # 状态校验
        if product["status"] not in ("draft", "active", "archived"):
            product["status"] = "draft"

        products.append(product)

    return products
