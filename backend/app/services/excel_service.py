"""Excel 导出服务: openpyxl 生成货盘表, 含图片嵌入"""
from pathlib import Path
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XlImage
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

from app.core.config import settings


# 样式常量
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
TITLE_FONT = Font(name="微软雅黑", size=16, bold=True)
CELL_FONT = Font(name="微软雅黑", size=10)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

# 列宽配置 (字符数)
COL_WIDTHS = {
    "A": 8,    # 序号
    "B": 20,   # 图片
    "C": 16,   # 款号
    "D": 20,   # 名称
    "E": 12,   # 品类
    "F": 10,   # 颜色
    "G": 16,   # 尺码
    "H": 10,   # 起订量
    "I": 12,   # 价格
    "J": 10,   # 状态
    "K": 20,   # 备注
}

# 图片单元格行高 (像素)
IMAGE_ROW_HEIGHT = 80
# 图片显示尺寸 (像素)
IMAGE_DISPLAY_SIZE = 70


def export_catalog_to_excel(
    catalog_name: str,
    items: list[dict],
    customer_name: str | None = None,
) -> bytes:
    """
    将货盘数据导出为 Excel, 返回文件 bytes.
    items: [{ sku_code, name, category, color, size_range, price, min_order_qty, stock_status, note, image_path }]
    """
    wb = Workbook()
    ws = wb.active
    ws.title = catalog_name[:31]  # Excel sheet 名最长 31 字符
    temp_files: list[str] = []  # 临时水印文件路径, save 后统一删除

    # 设置列宽
    for col, width in COL_WIDTHS.items():
        ws.column_dimensions[col].width = width

    # 标题行
    ws.merge_cells("A1:K1")
    title_cell = ws["A1"]
    title_cell.value = f"货盘表 - {catalog_name}"
    title_cell.font = TITLE_FONT
    title_cell.alignment = CENTER
    ws.row_dimensions[1].height = 30

    # 客户信息行
    if customer_name:
        ws.merge_cells("A2:K2")
        info_cell = ws["A2"]
        info_cell.value = f"客户: {customer_name}"
        info_cell.font = Font(name="微软雅黑", size=10)
        info_cell.alignment = Alignment(horizontal="left", vertical="center")

    # 表头行 (第3行, 或第2行如无客户)
    header_row = 3 if customer_name else 2
    headers = ["序号", "图片", "款号", "名称", "品类", "颜色", "尺码", "起订量", "价格", "状态", "备注"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    ws.row_dimensions[header_row].height = 25

    # 数据行
    status_map = {"available": "可供", "low_stock": "紧张", "sold_out": "断货"}
    for idx, item in enumerate(items, 1):
        row = header_row + idx
        ws.row_dimensions[row].height = IMAGE_ROW_HEIGHT

        values = [
            idx,
            None,  # 图片列, 下面单独插入
            item.get("sku_code", ""),
            item.get("name", ""),
            item.get("category", ""),
            item.get("color", ""),
            item.get("size_range", ""),
            item.get("min_order_qty", 1),
            item.get("price", 0),
            status_map.get(item.get("stock_status", ""), ""),
            item.get("note", ""),
        ]

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = CELL_FONT
            cell.alignment = CENTER
            cell.border = THIN_BORDER

        # 价格列格式
        price_cell = ws.cell(row=row, column=9)
        price_cell.number_format = "¥#,##0.00"

        # 插入图片 (带水印)
        image_path = item.get("image_path")
        if image_path:
            full_path = Path(settings.upload_dir) / image_path
            if full_path.exists():
                try:
                    import tempfile
                    from app.services.image_service import add_watermark
                    with tempfile.NamedTemporaryFile(suffix=".jpg", delete=False) as tmp:
                        tmp_path = tmp.name
                    add_watermark(full_path, tmp_path)
                    img = XlImage(tmp_path)
                    img.width = IMAGE_DISPLAY_SIZE
                    img.height = IMAGE_DISPLAY_SIZE
                    cell_ref = f"B{row}"
                    ws.add_image(img, cell_ref)
                    temp_files.append(tmp_path)  # save 后统一删除
                except Exception:
                    ws.cell(row=row, column=2, value="[图片]")

    # 冻结表头
    ws.freeze_panes = f"A{header_row + 1}"

    # 输出为 bytes
    buffer = BytesIO()
    wb.save(buffer)

    # 清理临时水印文件
    for tf in temp_files:
        Path(tf).unlink(missing_ok=True)

    return buffer.getvalue()
